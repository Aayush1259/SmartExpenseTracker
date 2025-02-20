import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from database import Database

class Export:
    def __init__(self, db: Database):
        self.db = db

    def to_csv(self, filename: str) -> bool:
        try:
            with open(filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["ID", "Date", "Amount", "Category", "Description"])
                writer.writerows(self.db.get_expenses())
            return True
        except Exception as e:
            print("CSV export error:", e)
            return False

    def to_excel(self, filename: str) -> bool:
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Expenses"
            headers = ["ID", "Date", "Amount", "Category", "Description"]
            ws.append(headers)
            for row in self.db.get_expenses():
                ws.append(row)

            data = pd.DataFrame(self.db.get_expenses(),
                                columns=["id", "date", "amount", "category", "description"])
            if not data.empty:
                data['date'] = pd.to_datetime(data['date'])
            else:
                data = pd.DataFrame(columns=["id", "date", "amount", "category", "description"])

            monthly = data.set_index('date').resample("M").sum().reset_index()
            monthly['date'] = monthly['date'].dt.strftime('%Y-%m')
            ws_monthly = wb.create_sheet("Monthly Summary")
            for r in dataframe_to_rows(monthly, index=False, header=True):
                ws_monthly.append(r)

            weekly = data.set_index('date').resample("W-Mon").sum().reset_index()
            weekly['date'] = weekly['date'].dt.strftime('%Y-%m-%d')
            ws_weekly = wb.create_sheet("Weekly Summary")
            for r in dataframe_to_rows(weekly, index=False, header=True):
                ws_weekly.append(r)

            yearly = data.set_index('date').resample("Y").sum().reset_index()
            yearly['date'] = yearly['date'].dt.year
            ws_yearly = wb.create_sheet("Yearly Summary")
            for r in dataframe_to_rows(yearly, index=False, header=True):
                ws_yearly.append(r)

            self._format_workbook(wb)
            wb.save(filename)
            return True
        except Exception as e:
            print("Excel export error:", e)
            return False

    def _format_workbook(self, wb: Workbook):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="32CD32", end_color="32CD32", fill_type="solid")
        alignment = Alignment(horizontal="center")
        for sheet in wb.worksheets:
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
            for col in sheet.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                sheet.column_dimensions[get_column_letter(col[0].column)].width = (max_length + 2) * 1.2
            if sheet.title == "Expenses":
                for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
                    for cell in row:
                        cell.number_format = '"$"#,##0.00'
